"""Montagem da lista completa (composição, inclusões manuais e pendências) para Excel/PDF."""

from __future__ import annotations

from decimal import Decimal
from html import escape
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

try:
    from reportlab.platypus.tables import LongTable
except ImportError:
    LongTable = Table

from composicao_painel.models import (
    ComposicaoInclusaoManual,
    ComposicaoItem,
    PendenciaItem,
)
from core.choices.cargas import TipoCargaChoices
from projetos.models import Projeto


COLUNAS = [
    "Origem",
    "Parte do painel",
    "Categoria",
    "Tag carga",
    "Descrição carga",
    "Tipo carga",
    "Potência",
    "Corrente ref. (A)",
    "Código produto",
    "Descrição produto",
    "Fabricante",
    "Quantidade",
    "Observações",
    "Memória de cálculo",
    "Status",
]


def _txt(v) -> str:
    if v is None:
        return ""
    if isinstance(v, Decimal):
        s = format(v, "f").rstrip("0").rstrip(".")
        return s or "0"
    s = str(v).strip()
    return s


def _potencia_carga(carga) -> str:
    if carga is None:
        return ""
    if str(carga.tipo) != TipoCargaChoices.MOTOR.value:
        return ""
    try:
        motor = carga.motor
    except Exception:
        return ""
    if motor is None or motor.potencia_corrente_valor is None:
        return ""
    unidade = (
        motor.get_potencia_corrente_unidade_display()
        if hasattr(motor, "get_potencia_corrente_unidade_display")
        else str(motor.potencia_corrente_unidade or "")
    )
    return f"{motor.potencia_corrente_valor} {unidade}".strip()


def _corrente_para_carga(corrente_referencia_a, carga) -> str:
    """Corrente de referência explícita ou derivada da carga (motor/resistência)."""
    if corrente_referencia_a is not None:
        return _txt(corrente_referencia_a)
    if carga is None:
        return ""
    if str(carga.tipo) == TipoCargaChoices.MOTOR.value:
        try:
            m = carga.motor
            if m and m.corrente_calculada_a is not None:
                return _txt(m.corrente_calculada_a)
        except Exception:
            pass
    if str(carga.tipo) == TipoCargaChoices.RESISTENCIA.value:
        try:
            r = carga.resistencia
            if r and r.corrente_calculada_a is not None:
                return _txt(r.corrente_calculada_a)
        except Exception:
            pass
    return ""


def _corrente_ref_item(item: ComposicaoItem) -> str:
    return _corrente_para_carga(item.corrente_referencia_a, item.carga)


def montar_linhas_export(projeto: Projeto) -> tuple[list[str], list[list[str]]]:
    """Retorna (cabeçalhos, linhas) com strings para planilha/PDF."""
    linhas: list[list[str]] = []

    qs_comp = (
        ComposicaoItem.objects.filter(projeto=projeto)
        .select_related(
            "produto",
            "produto__categoria",
            "carga",
            "carga__motor",
            "carga__resistencia",
        )
        .order_by("ordem", "id")
    )
    for item in qs_comp:
        carga = item.carga
        linhas.append(
            [
                "Composição aprovada",
                item.get_parte_painel_display(),
                item.get_categoria_produto_display(),
                _txt(carga.tag) if carga else "",
                _txt(carga.descricao) if carga else "",
                carga.get_tipo_display() if carga else "",
                _potencia_carga(carga),
                _corrente_ref_item(item),
                _txt(item.produto.codigo),
                _txt(item.produto.descricao),
                _txt(item.produto.fabricante),
                _txt(item.quantidade),
                _txt(item.observacoes),
                _txt(item.memoria_calculo),
                "",
            ]
        )

    qs_inc = (
        ComposicaoInclusaoManual.objects.filter(projeto=projeto)
        .select_related("produto", "produto__categoria")
        .order_by("ordem", "id")
    )
    for inc in qs_inc:
        p = inc.produto
        cat = p.categoria
        linhas.append(
            [
                "Inclusão manual (catálogo)",
                "—",
                cat.get_nome_display(),
                "",
                "",
                "",
                "",
                "",
                _txt(p.codigo),
                _txt(p.descricao),
                _txt(p.fabricante),
                _txt(inc.quantidade),
                _txt(inc.observacoes),
                "",
                "",
            ]
        )

    qs_pend = (
        PendenciaItem.objects.filter(projeto=projeto)
        .select_related(
            "carga",
            "carga__motor",
            "carga__resistencia",
        )
        .order_by("ordem", "id")
    )
    for pend in qs_pend:
        carga = pend.carga
        linhas.append(
            [
                "Pendência (catálogo)",
                pend.get_parte_painel_display(),
                pend.get_categoria_produto_display(),
                _txt(carga.tag) if carga else "",
                _txt(carga.descricao) if carga else "",
                carga.get_tipo_display() if carga else "",
                _potencia_carga(carga),
                _corrente_para_carga(pend.corrente_referencia_a, carga),
                "",
                _txt(pend.descricao),
                "",
                "",
                _txt(pend.observacoes),
                _txt(pend.memoria_calculo),
                pend.get_status_display(),
            ]
        )

    return COLUNAS, linhas


def nome_arquivo_seguro(projeto: Projeto, extensao: str) -> str:
    base = _txt(projeto.codigo) or str(projeto.id)[:8]
    safe = "".join(c if c.isalnum() or c in "-_" else "_" for c in base) or "projeto"
    return f"{safe}_composicao_completa.{extensao}"


def render_xlsx_bytes(projeto: Projeto, header: list[str], linhas: list[list[str]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Lista completa"

    bold = Font(bold=True)
    ws.append(["Projeto (código)", _txt(projeto.codigo)])
    ws.append(["Projeto (nome)", _txt(projeto.nome)])
    ws.append([])
    for c in range(1, 3):
        ws.cell(row=1, column=c).font = bold
        ws.cell(row=2, column=c).font = bold

    ws.append(header)
    for col in range(1, len(header) + 1):
        ws.cell(row=4, column=col).font = bold

    for row in linhas:
        ws.append(row)

    max_col = ws.max_column or len(header)
    for idx in range(1, max_col + 1):
        letter = get_column_letter(idx)
        vals = []
        for row in ws.iter_rows(min_col=idx, max_col=idx, values_only=True):
            vals.append(len(str(row[0] or "")))
        maxlen = min(48, max(vals) if vals else 10)
        ws.column_dimensions[letter].width = min(56, maxlen + 2)

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _pdf_para_text(raw: str) -> str:
    """Texto seguro para ReportLab Paragraph (quebra de linha + entidades HTML)."""
    if not raw or not str(raw).strip():
        return " "
    return escape(str(raw).strip()).replace("\n", "<br/>")


# Frações da largura útil (15 colunas); soma = 1 — colunas de texto recebem mais espaço.
_PDF_COL_FRACS = (
    0.060,
    0.070,
    0.065,
    0.038,
    0.085,
    0.042,
    0.046,
    0.041,
    0.050,
    0.125,
    0.055,
    0.032,
    0.085,
    0.115,
    0.091,
)


def render_pdf_bytes(projeto: Projeto, header: list[str], linhas: list[list[str]]) -> bytes:
    bio = BytesIO()
    page_w, _page_h = landscape(A4)
    margin = 1.1 * cm
    doc = SimpleDocTemplate(
        bio,
        pagesize=landscape(A4),
        leftMargin=margin,
        rightMargin=margin,
        topMargin=margin,
        bottomMargin=margin,
    )
    usable_w = page_w - doc.leftMargin - doc.rightMargin

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "pdf_title",
        parent=styles["Heading1"],
        fontSize=14,
        leading=17,
        spaceAfter=6,
    )
    title = Paragraph(
        f"<b>Composição completa do painel</b><br/>"
        f"<font size=\"10\" color=\"#333333\">{_pdf_para_text(_txt(projeto.codigo))} — "
        f"{_pdf_para_text(_txt(projeto.nome))}</font>",
        title_style,
    )

    style_cell = ParagraphStyle(
        "pdf_cell",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=6.5,
        leading=8,
        alignment=TA_LEFT,
        wordWrap="LTR",
    )
    style_header = ParagraphStyle(
        "pdf_th",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=6.5,
        leading=8,
        alignment=TA_CENTER,
        textColor=colors.white,
        wordWrap="LTR",
    )

    col_widths = [usable_w * f for f in _PDF_COL_FRACS]
    # Garantir que a soma bate com a largura útil (evita desalinhamento por arredondamento).
    drift = usable_w - sum(col_widths)
    col_widths[-1] = max(col_widths[-1] + drift, 20)

    body_rows: list[list[str]] = [list(r) for r in linhas]
    if not body_rows:
        body_rows = [
            ["(sem itens na composição, inclusões manuais ou pendências)"]
            + [""] * (len(header) - 1)
        ]

    table_data: list = [
        [Paragraph(_pdf_para_text(h), style_header) for h in header],
    ]
    for row in body_rows:
        table_data.append([Paragraph(_pdf_para_text(c), style_cell) for c in row])

    tbl = LongTable(table_data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2c5282")),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e0")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                ("LEFTPADDING", (0, 0), (-1, -1), 3),
                ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f7fafc")]),
            ]
        )
    )

    story = [title, Spacer(1, 0.25 * cm), tbl]
    doc.build(story)
    return bio.getvalue()
